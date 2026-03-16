"""Document loaders for PDF, Markdown, plain text, and spreadsheets."""

import csv
from pathlib import Path
from typing import List

from langchain_core.documents import Document

from docops.ingestion.metadata import build_doc_id
from docops.logging import get_logger

logger = get_logger("docops.ingestion.loaders")


def _tabular_row_to_text(headers: list[str], values: list[str], row_idx: int) -> str:
    pairs = []
    width = max(len(headers), len(values))
    for idx in range(width):
        key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
        value = values[idx] if idx < len(values) else ""
        key = (key or f"col_{idx + 1}").strip()
        value = (value or "").strip()
        pairs.append(f"{key}: {value}")
    return f"Linha {row_idx}: " + " | ".join(pairs)


def _build_tabular_content(
    *,
    title: str,
    headers: list[str],
    rows: list[list[str]],
    max_rows: int = 2000,
) -> str:
    normalized_headers = [(header or "").strip() for header in headers]
    normalized_headers = [
        header if header else f"col_{idx + 1}"
        for idx, header in enumerate(normalized_headers)
    ]
    content_lines = [title]
    if normalized_headers:
        content_lines.append("Colunas: " + " | ".join(normalized_headers))
    content_lines.append("")

    total_rows = len(rows)
    visible_rows = rows[:max_rows]
    for row_idx, row in enumerate(visible_rows, start=1):
        content_lines.append(_tabular_row_to_text(normalized_headers, row, row_idx))

    if total_rows > max_rows:
        content_lines.append("")
        content_lines.append(
            f"[Aviso] Apenas as primeiras {max_rows} linhas foram ingeridas de um total de {total_rows}."
        )
    return "\n".join(content_lines)


def load_pdf(file_path: Path) -> List[Document]:
    """Load a PDF file, one Document per page, with page metadata."""
    try:
        import pypdf
    except ImportError:
        raise ImportError("pypdf is required for PDF loading: pip install pypdf")

    logger.info(f"Loading PDF: {file_path}")
    source_path = str(file_path)
    doc_id = build_doc_id(source_path)
    docs = []

    with open(file_path, "rb") as f:
        reader = pypdf.PdfReader(f)
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            page_no = i + 1
            docs.append(Document(
                page_content=text,
                metadata={
                    "file_name": file_path.name,
                    "source": source_path,
                    "source_path": source_path,
                    "doc_id": doc_id,
                    "file_type": "pdf",
                    "page": page_no,
                    "page_start": page_no,
                    "page_end": page_no,
                    "section_title": "",
                    "section_path": "",
                },
            ))

    logger.info(f"Loaded {len(docs)} pages from {file_path.name}")
    return docs


def load_text(file_path: Path) -> List[Document]:
    """Load a plain text (.txt) file as a single Document."""
    logger.info(f"Loading TXT: {file_path}")
    content = file_path.read_text(encoding="utf-8", errors="replace")
    source_path = str(file_path)
    doc = Document(
        page_content=content,
        metadata={
            "file_name": file_path.name,
            "source": source_path,
            "source_path": source_path,
            "doc_id": build_doc_id(source_path),
            "file_type": "txt",
            "page": "N/A",
            "page_start": "N/A",
            "page_end": "N/A",
            "section_title": "",
            "section_path": "",
        },
    )
    return [doc]


def load_markdown(file_path: Path) -> List[Document]:
    """Load a Markdown file as a single Document."""
    logger.info(f"Loading MD: {file_path}")
    content = file_path.read_text(encoding="utf-8", errors="replace")
    source_path = str(file_path)
    doc = Document(
        page_content=content,
        metadata={
            "file_name": file_path.name,
            "source": source_path,
            "source_path": source_path,
            "doc_id": build_doc_id(source_path),
            "file_type": "md",
            "page": "N/A",
            "page_start": "N/A",
            "page_end": "N/A",
            "section_title": "",
            "section_path": "",
        },
    )
    return [doc]


def load_csv(file_path: Path) -> List[Document]:
    """Load a CSV file as one normalized tabular Document."""
    logger.info(f"Loading CSV: {file_path}")
    source_path = str(file_path)
    with open(file_path, "r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        raw_rows = [[str(cell) for cell in row] for row in reader]

    if not raw_rows:
        headers: list[str] = []
        rows: list[list[str]] = []
    else:
        headers = raw_rows[0]
        rows = raw_rows[1:]

    content = _build_tabular_content(
        title=f"Tabela CSV: {file_path.name}",
        headers=headers,
        rows=rows,
    )
    doc = Document(
        page_content=content,
        metadata={
            "file_name": file_path.name,
            "source": source_path,
            "source_path": source_path,
            "doc_id": build_doc_id(source_path),
            "file_type": "csv",
            "page": "N/A",
            "page_start": "N/A",
            "page_end": "N/A",
            "section_title": "",
            "section_path": "",
        },
    )
    return [doc]


def load_xlsx(file_path: Path) -> List[Document]:
    """Load an XLSX workbook with one Document per sheet."""
    logger.info(f"Loading XLSX: {file_path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX loading: pip install openpyxl"
        ) from exc

    source_path = str(file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    docs: list[Document] = []

    for sheet_idx, sheet in enumerate(wb.worksheets, start=1):
        rows_iter = sheet.iter_rows(values_only=True)
        raw_rows = []
        for row in rows_iter:
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                raw_rows.append(values)

        headers = raw_rows[0] if raw_rows else []
        rows = raw_rows[1:] if len(raw_rows) > 1 else []
        content = _build_tabular_content(
            title=f"Planilha XLSX: {file_path.name} | Aba: {sheet.title}",
            headers=headers,
            rows=rows,
        )

        docs.append(
            Document(
                page_content=content,
                metadata={
                    "file_name": file_path.name,
                    "source": source_path,
                    "source_path": source_path,
                    "doc_id": build_doc_id(source_path),
                    "file_type": "xlsx",
                    "page": sheet_idx,
                    "page_start": sheet_idx,
                    "page_end": sheet_idx,
                    "section_title": sheet.title,
                    "section_path": sheet.title,
                },
            )
        )

    wb.close()
    return docs


_LOADERS = {
    ".pdf": load_pdf,
    ".txt": load_text,
    ".md": load_markdown,
    ".markdown": load_markdown,
    ".csv": load_csv,
    ".xlsx": load_xlsx,
}

SUPPORTED_EXTENSIONS = set(_LOADERS.keys())


def load_file(file_path: Path) -> List[Document]:
    """Dispatch to the correct loader based on file extension."""
    ext = file_path.suffix.lower()
    loader_fn = _LOADERS.get(ext)
    if loader_fn is None:
        raise ValueError(
            f"Unsupported file type '{ext}'. "
            f"Supported: {sorted(SUPPORTED_EXTENSIONS)}"
        )
    return loader_fn(file_path)


def load_directory(dir_path: Path) -> List[Document]:
    """Load all supported documents from a directory (non-recursive)."""
    dir_path = Path(dir_path)
    if not dir_path.exists():
        raise FileNotFoundError(f"Directory not found: {dir_path}")

    all_docs: List[Document] = []
    found_files = [
        f for f in sorted(dir_path.iterdir())
        if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
    ]

    if not found_files:
        logger.warning(
            f"No supported files found in '{dir_path}'. "
            f"Add PDF, MD, TXT, CSV, or XLSX files and run ingest again."
        )
        return []

    for file_path in found_files:
        try:
            docs = load_file(file_path)
            all_docs.extend(docs)
        except Exception as exc:
            logger.error(f"Failed to load '{file_path.name}': {exc}")

    logger.info(f"Loaded {len(all_docs)} document chunks from {len(found_files)} files")
    return all_docs
